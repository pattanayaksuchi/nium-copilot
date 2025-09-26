"""Build corridor- and payout-method-specific JSON Schemas from the validation workbook."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from app import settings


class DebugOptions:
    def __init__(self) -> None:
        self.enabled = False
        self.country: Optional[str] = None
        self.currency: Optional[str] = None
        self.method: Optional[str] = None
        self.channel: Optional[str] = None

    def configure(
        self,
        *,
        country: Optional[str] = None,
        currency: Optional[str] = None,
        method: Optional[str] = None,
        channel: Optional[str] = None,
    ) -> None:
        if country or currency or method or channel:
            self.enabled = True
        self.country = country.upper() if country else None
        self.currency = currency.upper() if currency else None
        self.method = method.lower() if method else None
        self.channel = channel.lower() if channel else None

    def match(self, *, country: str, currency: str, method: str, channel: str) -> bool:
        if not self.enabled:
            return False
        if self.country and self.country != country.upper():
            return False
        if self.currency and self.currency != currency.upper():
            return False
        if self.method and self.method != method.lower():
            return False
        if self.channel and self.channel != channel.lower():
            return False
        return True


DEBUG = DebugOptions()

EXPECTED_COLUMN_ORDER = [
    "Country",
    "Currency",
    "Field",
    "Mandatory",
    "Data Type",
    "Length",
    "Regex",
    "Allowed Values",
    "Conditional",
    "Notes",
]
EXPECTED_COLUMNS = set(EXPECTED_COLUMN_ORDER)

COUNTRY_CURRENCY_RE = re.compile(r"^(?P<country>.+?)\s*\((?P<currency>[A-Z0-9]{3})\)$")
COUNTRY_CURRENCY_DESCRIPTOR_RE = re.compile(
    r"^(?P<country>.+?)\s*\((?P<currency>[A-Z0-9]{3})\)\s*(?P<descriptor>.*)$"
)
MATRIX_HEADER_ROWS = [1, 2]
MANDATORY_TOKEN = re.compile(r"mandatory|required", re.IGNORECASE)
AFFIRMATIVE_TOKEN = re.compile(r"^(yes|y|true)$", re.IGNORECASE)
CONDITIONAL_TOKEN = re.compile(r"conditional", re.IGNORECASE)
CODE_LIKE_RE = re.compile(r"^[A-Za-z0-9_.]+$")
STATE_ALIASES = {
    "mandatory": "mandatory",
    "required": "mandatory",
    "conditional": "conditional",
    "optional": "optional",
}
CONDITIONAL_PHRASES = (
    "required if",
    "required when",
    "required for",
    "required unless",
    "mandatory if",
    "mandatory when",
    "mandatory for",
    "mandatory unless",
    "conditionally",
)
MUST_MANDATORY_PHRASES = (
    "must provide",
    "must include",
    "must submit",
    "must supply",
    "must have",
    "must upload",
    "must enter",
    "must quote",
    "must list",
)

SHEET_METHOD_MAP = {
    "bank": "bank",
    "wallet": "wallet",
    "proxy": "proxy",
    "card": "card",
    "cash": "cash",
}
DEFAULT_CHANNEL_BY_METHOD = {
    "bank": "local",
    "wallet": "default",
    "proxy": "default",
    "card": "default",
    "cash": "default",
}
COLOR_STATE_MAP: Dict[str, str] = {}
WIRE_HINTS = {
    "wire",
    "swift",
    "international",
    "cross border",
    "cross-border",
    "crossborder",
    "xb",
    "global",
    "intl",
    "overseas",
}
LOCAL_HINTS = {
    "local",
    "domestic",
    "ach",
    "sepa",
    "rtgs",
    "neft",
    "in-country",
    "incountry",
}


def clean_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, str):
        text = value
    else:
        text = str(value)
    normalized = (
        text.replace("\u00a0", " ")
        .replace("\r\n", " ")
        .replace("\n", " ")
        .replace("\t", " ")
        .strip()
    )
    if normalized.lower() == "nan":
        return ""
    return normalized


def infer_method(sheet_name: str) -> Optional[str]:
    name = clean_text(sheet_name).lower()
    if not name:
        return None
    if "instruction" in name:
        return None
    for keyword, method in SHEET_METHOD_MAP.items():
        if keyword in name:
            return method
    return None


def build_color_state_map(workbook_path: Path) -> None:
    COLOR_STATE_MAP.clear()
    try:
        wb = openpyxl_load_workbook(workbook_path, data_only=False)
    except ValueError:
        return
    instructions_sheet: Optional[Worksheet] = None
    for sheet_name in wb.sheetnames:
        if "instruction" in sheet_name.lower():
            instructions_sheet = wb[sheet_name]
            break
    if instructions_sheet is None:
        wb.close()
        return

    for row in (2, 3, 4):
        color_cell = instructions_sheet.cell(row=row, column=2)
        label_cell = instructions_sheet.cell(row=row, column=3)
        label = clean_text(label_cell.value).lower()
        state = STATE_ALIASES.get(label)
        if not state:
            continue
        key = color_key(color_cell.fill.start_color)
        if not key:
            key = color_key(color_cell.fill.fgColor)
        if key:
            COLOR_STATE_MAP[key] = state
    wb.close()

def normalize_color_code(value: Any) -> str:
    if value is None:
        return ""
    # Handle openpyxl Color objects and related wrappers.
    if hasattr(value, "rgb") and getattr(value, "rgb"):
        value = value.rgb
    elif hasattr(value, "value") and getattr(value, "value"):
        value = value.value
    elif hasattr(value, "indexed") and getattr(value, "indexed") is not None:
        indexed = getattr(value, "indexed")
        return f"INDEXED:{indexed}"

    if isinstance(value, bytes):
        try:
            value = value.decode("utf-8")
        except Exception:
            value = value.decode("latin1", errors="ignore")

    code = str(value).strip().upper()
    if not code:
        return ""
    if len(code) == 6:
        return f"FF{code}"
    if len(code) == 8:
        return code
    return ""


def color_key(color: Any) -> str:
    if color is None:
        return ""
    rgb = getattr(color, "rgb", None)
    if rgb:
        normalized = normalize_color_code(rgb)
        if normalized:
            return f"rgb:{normalized}"
    theme = getattr(color, "theme", None)
    tint = getattr(color, "tint", None)
    if theme is not None:
        return f"theme:{theme}:{tint or 0}"
    indexed = getattr(color, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"
    type_name = getattr(color, "type", None)
    if type_name:
        return f"{type_name}:{getattr(color, 'value', '')}"
    return ""


def classify_rgb(rgb: str) -> str:
    if not rgb:
        return "optional"
    hex_part = normalize_color_code(rgb)[-6:]
    if len(hex_part) != 6:
        return "optional"
    r = int(hex_part[0:2], 16)
    g = int(hex_part[2:4], 16)
    b = int(hex_part[4:6], 16)
    if g > r + 25 and g > b + 25:
        return "mandatory"
    if r >= g and r >= b:
        return "conditional"
    return "optional"


def mandatory_state_from_color(cell: Optional[Cell]) -> str:
    if cell is None:
        return "optional"
    fill = getattr(cell, "fill", None)
    if fill is None or fill.fill_type != "solid":
        if DEBUG.enabled:
            print(
                f"[DEBUG] fill check: coord={cell.coordinate if cell else None} "
                f"fill_type={getattr(fill, 'fill_type', None)}"
            )
        return "optional"
    key = color_key(fill.start_color)
    if not key:
        key = color_key(getattr(fill, "fgColor", None))
    if DEBUG.enabled:
        print(
            f"[DEBUG] color key={key} coord={cell.coordinate} fill_type={fill.fill_type}"
        )
    mapped = COLOR_STATE_MAP.get(key)
    if mapped:
        return mapped
    rgb = normalize_color_code(getattr(fill.start_color, "rgb", None))
    return classify_rgb(rgb)


def mandatory_state_from_text(value: Any) -> str:
    text = clean_text(value).lower()
    if not text:
        return ""
    if "optional" in text:
        return "optional"
    if any(phrase in text for phrase in CONDITIONAL_PHRASES):
        return "conditional"
    if any(phrase in text for phrase in MUST_MANDATORY_PHRASES):
        return "mandatory"
    if AFFIRMATIVE_TOKEN.match(text) or MANDATORY_TOKEN.search(text):
        return "mandatory"
    if text.startswith("conditional"):
        return "conditional"
    return ""


def parse_corridor_label(label: str) -> Tuple[str, str, str]:
    text = clean_text(label)
    descriptor = ""
    descriptor_match = COUNTRY_CURRENCY_DESCRIPTOR_RE.match(text)
    if descriptor_match:
        country = clean_text(descriptor_match.group("country"))
        currency = clean_text(descriptor_match.group("currency"))
        descriptor = clean_text(descriptor_match.group("descriptor"))
        return country, currency, descriptor

    for delimiter in (" - ", " – "):
        if delimiter in text:
            left, right = text.split(delimiter, 1)
            descriptor = clean_text(right)
            text = left
            break
    match = COUNTRY_CURRENCY_RE.match(text.strip())
    if match:
        country = clean_text(match.group("country"))
        currency = clean_text(match.group("currency"))
    else:
        country = clean_text(text)
        currency = ""
    return country, currency, descriptor


def infer_channel(method: str, corridor_label: str, descriptor: str) -> str:
    method_key = clean_text(method).lower() or "bank"
    default_channel = DEFAULT_CHANNEL_BY_METHOD.get(method_key, "default")
    if method_key != "bank":
        return default_channel
    text = f"{descriptor} {corridor_label}".lower()
    if "wire" in text or "swift" in text or "international" in text:
        return "wire"
    if "ach" in text or "local" in text or "domestic" in text:
        return "local"
    return default_channel


def first_non_empty_column(df: pd.DataFrame) -> Optional[pd.Series]:
    for column in df.columns:
        series = df[column]
        if series.dropna().empty:
            continue
        return series
    return None


def annotate_mandatory_states(df: pd.DataFrame, worksheet: Optional[Worksheet]) -> pd.DataFrame:
    if df.empty:
        return df
    result = df.copy()
    if "Mandatory" not in result.columns:
        result["Mandatory"] = ""

    color_states: List[str] = []
    if worksheet is not None and "Field" in result.columns:
        try:
            field_position = list(result.columns).index("Field") + 1
        except ValueError:
            field_position = None
        if field_position is not None:
            for offset, _ in enumerate(result.itertuples(index=False), start=2):
                cell = worksheet.cell(row=offset, column=field_position)
                color_states.append(mandatory_state_from_color(cell))
    if not color_states:
        color_states = ["optional" for _ in range(len(result))]
    if len(color_states) != len(result):
        fill_value = ["optional"] * len(result)
        for idx, state in enumerate(color_states[: len(result)]):
            fill_value[idx] = state
        color_states = fill_value

    result["__mandatory_state__"] = color_states

    textual_states = result["Mandatory"].apply(mandatory_state_from_text)
    mask = result["__mandatory_state__"].isin(["optional", ""]) & textual_states.astype(bool)
    result.loc[mask, "__mandatory_state__"] = textual_states[mask]

    result.loc[result["__mandatory_state__"] == "mandatory", "Mandatory"] = "mandatory"
    result.loc[result["__mandatory_state__"] == "conditional", "Mandatory"] = "conditional"
    return result


def extract_matrix_records(
    sheet_name: str,
    sheet_df: pd.DataFrame,
    method: str,
    worksheet: Optional[Worksheet],
) -> pd.DataFrame:
    if sheet_df.empty or not isinstance(sheet_df.columns, pd.MultiIndex):
        return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)

    base_columns = set(sheet_df.columns.get_level_values(0))
    if "Field name" not in base_columns:
        return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)

    column_lookup = {column: position for position, column in enumerate(sheet_df.columns, start=1)}

    field_block = sheet_df.xs("Field name", axis=1, level=0)
    field_series = first_non_empty_column(field_block)
    if field_series is None:
        return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)
    field_series = field_series.apply(clean_text)

    field_column_index = column_lookup.get(("Field name", field_block.columns[0]))
    row_lookup: Dict[str, int] = {}
    if worksheet is not None and field_column_index is not None:
        max_row = worksheet.max_row
        for row_number in range(1, max_row + 1):
            raw_value = worksheet.cell(row=row_number, column=field_column_index).value
            key = clean_text(raw_value).lower()
            if key and key not in row_lookup:
                row_lookup[key] = row_number

    alias_series: Optional[pd.Series] = None
    if "N1 Fields Name" in base_columns:
        alias_block = sheet_df.xs("N1 Fields Name", axis=1, level=0)
        alias_candidate = first_non_empty_column(alias_block)
        if alias_candidate is not None:
            alias_series = alias_candidate.apply(clean_text)

    corridor_names: List[str] = [
        name
        for name in sheet_df.columns.get_level_values(0).unique()
        if name not in {"Field name", "N1 Fields Name"}
    ]
    records: List[Dict[str, Any]] = []

    def derive_mandatory_and_conditional(texts: Sequence[str]) -> Tuple[str, str]:
        mandatory_value = ""
        conditional_value = ""
        for text in texts:
            normalized = clean_text(text)
            if not normalized:
                continue
            state_hint = mandatory_state_from_text(normalized)
            if state_hint == "conditional":
                mandatory_value = "conditional"
                if not conditional_value:
                    conditional_value = normalized
                continue
            lower = normalized.lower()
            if not mandatory_value and (state_hint == "mandatory" or MANDATORY_TOKEN.search(lower) or AFFIRMATIVE_TOKEN.match(lower)):
                mandatory_value = "mandatory"
        return mandatory_value, conditional_value

    for row_position, (idx, field_name) in enumerate(field_series.items()):
        normalized_field = clean_text(field_name)
        if not normalized_field or normalized_field.lower() == "notes":
            continue
        row_lookup_key = normalized_field.lower()
        canonical_field = normalized_field
        if alias_series is not None:
            alias_value = clean_text(alias_series.get(idx, ""))
            if alias_value and CODE_LIKE_RE.match(alias_value):
                canonical_field = alias_value

        excel_row = row_lookup.get(row_lookup_key)
        if excel_row is None and MATRIX_HEADER_ROWS:
            header_offset = (max(MATRIX_HEADER_ROWS) + 2) if MATRIX_HEADER_ROWS else 1
            excel_row = row_position + header_offset

        for corridor in corridor_names:
            corridor_block = sheet_df.xs(corridor, axis=1, level=0)
            if idx not in corridor_block.index:
                continue
            row_values = corridor_block.loc[idx]
            if isinstance(row_values, pd.Series):
                cleaned = {key: clean_text(value) for key, value in row_values.items()}
            else:
                cleaned = {corridor_block.columns: clean_text(row_values)}  # type: ignore[assignment]
            non_empty_items = {key: value for key, value in cleaned.items() if value}
            if not non_empty_items:
                continue

            country_name, currency_code, descriptor = parse_corridor_label(corridor)
            base_channel = infer_channel(method, corridor, descriptor)

            if DEBUG.enabled:
                print(
                    f"[DEBUG] non_empty_keys for {corridor}: {list(non_empty_items.keys())}"
                )

            channel_buckets: Dict[str, Dict[str, Any]] = {}

            def resolve_channel(raw_key: str) -> str:
                key = clean_text(raw_key)
                key_lower = key.lower()
                normalized = key_lower.replace("-", " ")
                if any(hint in normalized for hint in WIRE_HINTS):
                    return "wire"
                if any(hint in normalized for hint in LOCAL_HINTS):
                    return "local"
                if "." in key_lower:
                    suffix = key_lower.split(".")[-1]
                    if suffix.isdigit() and int(suffix) % 2 == 1:
                        return "wire"
                if "regex(wires" in key_lower or "description.1" in key_lower or "description.3" in key_lower:
                    return "wire"
                if "regex(local" in key_lower:
                    return "local"
                return base_channel

            for raw_key, value in non_empty_items.items():
                channel = resolve_channel(raw_key)
                bucket = channel_buckets.setdefault(
                    channel,
                    {
                        "descriptions": [],
                        "notes": [],
                        "regex": "",
                        "explicit_state": "",
                        "explicit_note": "",
                        "color_state": "",
                        "non_empty_keys": [],
                    },
                )
                bucket["non_empty_keys"].append(raw_key)

                key_text = clean_text(raw_key)
                lower_key = key_text.lower()
                if "regex" in lower_key and "description" not in lower_key:
                    if not bucket["regex"]:
                        bucket["regex"] = value
                    else:
                        bucket["notes"].append(f"{key_text}: {value}")
                elif "description" in lower_key:
                    bucket["descriptions"].append(value)
                else:
                    bucket["notes"].append(f"{key_text}: {value}")

                if worksheet is not None and excel_row is not None:
                    column_index = column_lookup.get((corridor, raw_key))
                    if column_index is not None:
                        cell = worksheet.cell(row=excel_row, column=column_index)
                        color_state = mandatory_state_from_color(cell)
                        if bucket["color_state"] != "mandatory" and color_state != "optional":
                            bucket["color_state"] = color_state
                        if color_state == "conditional" and bucket["explicit_state"] != "mandatory":
                            bucket["explicit_state"] = "conditional"
                            if not bucket["explicit_note"]:
                                bucket["explicit_note"] = "Conditionally mandatory (highlighted)."

                requirement_state = mandatory_state_from_text(value)
                bucket_explicit = bucket.get("explicit_state")
                if requirement_state != "mandatory" and bucket_explicit != "conditional":
                    if "must" in clean_text(value).lower():
                        requirement_state = "conditional"
                if requirement_state == "conditional":
                    bucket["explicit_state"] = "conditional"
                    if not bucket["explicit_note"]:
                        bucket["explicit_note"] = clean_text(value)
                elif requirement_state == "mandatory" and bucket["explicit_state"] != "conditional":
                    bucket["explicit_state"] = "mandatory"

            # incorporate row-level color if present
            row_color_state = ""
            if (
                worksheet is not None
                and excel_row is not None
                and field_column_index is not None
            ):
                row_color_state = mandatory_state_from_color(
                    worksheet.cell(row=excel_row, column=field_column_index)
                )

            total_channels = len(channel_buckets)

            for channel, bucket in channel_buckets.items():
                color_state_detected = bucket["color_state"] or row_color_state
                descriptions = bucket["descriptions"]
                notes_list = bucket["notes"]
                regex_value = bucket["regex"]
                explicit_state = bucket["explicit_state"] or row_color_state
                explicit_note = bucket["explicit_note"]

                derived_mandatory, derived_conditional = derive_mandatory_and_conditional(
                    descriptions + notes_list
                )
                if color_state_detected == "mandatory" and explicit_state != "mandatory":
                    explicit_state = "mandatory"
                if explicit_state == "conditional":
                    mandatory_value = "conditional"
                    conditional_value = explicit_note or derived_conditional
                elif explicit_state == "mandatory":
                    mandatory_value = "mandatory"
                    conditional_value = derived_conditional
                elif color_state_detected in {"mandatory", "conditional"}:
                    if color_state_detected == "mandatory":
                        mandatory_value = "mandatory"
                        conditional_value = derived_conditional
                    elif color_state_detected == "conditional" and (
                        channel != base_channel or derived_conditional or explicit_note
                    ):
                        mandatory_value = "conditional"
                        conditional_value = explicit_note or derived_conditional
                    else:
                        mandatory_value = derived_mandatory
                        conditional_value = derived_conditional
                else:
                    mandatory_value = derived_mandatory
                    conditional_value = derived_conditional

                notes_parts = descriptions + notes_list
                notes_text = " | ".join(part for part in notes_parts if part)

                record = {
                    "Country": country_name.upper(),
                    "Currency": currency_code.upper(),
                    "Field": canonical_field,
                    "Mandatory": mandatory_value,
                    "Data Type": "",
                    "Length": "",
                    "Regex": regex_value,
                    "Allowed Values": "",
                    "Conditional": conditional_value,
                    "Notes": notes_text or sheet_name,
                    "Method": method,
                    "Channel": channel,
                    "__mandatory_state__": mandatory_value or "optional",
                }

                if DEBUG.match(
                    country=record["Country"],
                    currency=record["Currency"],
                    method=record["Method"],
                    channel=record["Channel"],
                ):
                    print(
                        f"[DEBUG] {record['Country']}/{record['Currency']} {record['Method']}:{record['Channel']} "
                        f"field={record['Field']} mandatory={mandatory_value} conditional={conditional_value} "
                        f"explicit_state={explicit_state} color_state={color_state_detected} "
                        f"row_color={row_color_state} notes={notes_text}"
                    )
                if DEBUG.enabled and canonical_field == "remitter_postcode" and record["Channel"] == "local":
                    print(
                        f"[TRACE] remitter_postcode local -> explicit_state={explicit_state} "
                        f"color_state={color_state_detected} derived=({derived_mandatory}, {derived_conditional})"
                    )

                records.append(record)

    if not records:
        return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)
    return pd.DataFrame.from_records(records)


def normalize_sheet(
    workbook: pd.ExcelFile,
    sheet_name: str,
    base_df: pd.DataFrame,
    worksheet: Optional[Worksheet],
    method: str,
) -> pd.DataFrame:
    if base_df.empty:
        return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)

    if EXPECTED_COLUMNS.issubset(set(base_df.columns)):
        working = annotate_mandatory_states(base_df.copy(), worksheet)
    else:
        try:
            matrix_df = pd.read_excel(
                workbook,
                sheet_name=sheet_name,
                header=[1, 2],
                engine=workbook.engine,
            )
        except ValueError:
            return pd.DataFrame(columns=EXPECTED_COLUMN_ORDER)
        return extract_matrix_records(sheet_name, matrix_df, method, worksheet)

    default_channel = DEFAULT_CHANNEL_BY_METHOD.get(method, "default")
    if "Method" not in working.columns:
        working["Method"] = method
    working["Method"] = working["Method"].fillna(method).apply(lambda value: clean_text(value).lower() or method)

    if "Channel" not in working.columns:
        working["Channel"] = default_channel
    working["Channel"] = working["Channel"].apply(lambda value: clean_text(value).lower() or default_channel)

    if "__mandatory_state__" not in working.columns:
        working["__mandatory_state__"] = (
            working.get("Mandatory", "").apply(mandatory_state_from_text).replace({"": "optional"})
        )

    for column in EXPECTED_COLUMN_ORDER:
        if column not in working.columns:
            working[column] = ""

    return working


def load_workbook(path: Path) -> pd.ExcelFile:
    try:
        return pd.ExcelFile(path)
    except ValueError as exc:
        message = str(exc)
        if "could not read stylesheet" in message or "Max value is" in message:
            try:
                return pd.ExcelFile(path, engine="calamine")
            except ImportError as calamine_exc:  # pragma: no cover - guardrail
                raise RuntimeError(
                    "Unable to read workbook because its styles are corrupted. "
                    "Install the optional dependency `python-calamine` or provide a cleaned copy "
                    "of the workbook."
                ) from calamine_exc
        raise


def parse_length(value: Any) -> Tuple[int | None, int | None]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, None
    text = str(value).strip()
    if not text:
        return None, None
    if "-" in text:
        parts = text.replace(" ", "").split("-")
        try:
            return int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            return None, None
    try:
        length = int(float(text))
        return None, length
    except ValueError:
        return None, None


def data_type_to_schema(data_type: Any) -> Dict[str, Any]:
    mapping = {
        "string": "string",
        "str": "string",
        "text": "string",
        "number": "number",
        "float": "number",
        "decimal": "number",
        "integer": "integer",
        "int": "integer",
        "boolean": "boolean",
        "bool": "boolean",
        "array": "array",
        "list": "array",
        "object": "object",
        "dict": "object",
    }
    if data_type is None or (isinstance(data_type, float) and pd.isna(data_type)):
        return {"type": "string"}
    normalized = str(data_type).strip().lower()
    json_type = mapping.get(normalized, "string")
    return {"type": json_type}


def ensure_object_node(node: Dict[str, Any], key: str) -> Dict[str, Any]:
    properties = node.setdefault("properties", {})
    if key not in properties:
        properties[key] = {"type": "object", "properties": {}, "required": [], "additionalProperties": True}
    child = properties[key]
    child.setdefault("type", "object")
    child.setdefault("properties", {})
    child.setdefault("required", [])
    child.setdefault("additionalProperties", True)
    return child


def apply_field_schema(node: Dict[str, Any], field_parts: List[str], row: pd.Series) -> None:
    cursor = node
    for part in field_parts[:-1]:
        cursor = ensure_object_node(cursor, part)
    leaf_key = field_parts[-1]
    cursor.setdefault("properties", {})
    field_schema: Dict[str, Any] = data_type_to_schema(row.get("Data Type"))

    min_length, max_length = parse_length(row.get("Length"))
    if field_schema["type"] == "string":
        if min_length:
            field_schema["minLength"] = min_length
        if max_length:
            field_schema["maxLength"] = max_length
    if row.get("Regex"):
        field_schema["pattern"] = str(row["Regex"]).strip()
    if row.get("Allowed Values") and not pd.isna(row.get("Allowed Values")):
        raw_values = str(row["Allowed Values"]).strip()
        if raw_values:
            choices = [value.strip() for value in raw_values.replace("|", ",").split(",") if value.strip()]
            if choices:
                field_schema["enum"] = choices

    notes: List[str] = []
    if row.get("Conditional") and not pd.isna(row.get("Conditional")):
        notes.append(f"Conditional: {row['Conditional']}")
    if row.get("Notes") and not pd.isna(row.get("Notes")):
        notes.append(str(row["Notes"]).strip())

    cursor["properties"][leaf_key] = field_schema
    mandatory_state = str(row.get("__mandatory_state__", "")).strip().lower()
    if not mandatory_state:
        mandatory_state = mandatory_state_from_text(row.get("Mandatory"))
    if mandatory_state == "conditional":
        notes.append("Status: Conditionally mandatory (orange highlight).")
    if notes:
        existing_description = field_schema.get("description")
        combined_description = " | ".join(note for note in notes if note)
        if existing_description:
            field_schema["description"] = f"{existing_description} | {combined_description}"
        else:
            field_schema["description"] = combined_description
    if mandatory_state == "mandatory":
        cursor.setdefault("required", [])
        if leaf_key not in cursor["required"]:
            cursor["required"].append(leaf_key)


def build_schemas(df: pd.DataFrame) -> Dict[Tuple[str, str], Dict[str, Any]]:
    corridors: Dict[Tuple[str, str], Dict[str, Dict[str, Dict[str, Any]]]] = defaultdict(lambda: defaultdict(dict))

    for _, row in df.iterrows():
        field_value = clean_text(row.get("Field", ""))
        if not field_value:
            continue
        currency = clean_text(row.get("Currency", "")).upper()
        country = clean_text(row.get("Country", "")).upper()
        if not currency or not country:
            continue
        method = clean_text(row.get("Method", "bank")).lower() or "bank"
        channel = clean_text(row.get("Channel", "")).lower()
        if not channel:
            channel = DEFAULT_CHANNEL_BY_METHOD.get(method, "default")
        if method == "bank" and channel == "default":
            channel = "local"

        field_path = [part.strip() for part in field_value.split(".") if part.strip()]
        if not field_path:
            continue

        method_map = corridors[(currency, country)]
        channel_map = method_map.setdefault(method, {})
        channel_schema = channel_map.setdefault(
            channel,
            {"type": "object", "properties": {}, "required": [], "additionalProperties": True},
        )
        apply_field_schema(channel_schema, field_path, row)

    final: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for (currency, country), method_map in corridors.items():
        payout_methods: Dict[str, Any] = {}
        for method, channels in method_map.items():
            ordered_channels = dict(sorted(channels.items()))
            default_channel = DEFAULT_CHANNEL_BY_METHOD.get(method, next(iter(ordered_channels), "default"))
            if default_channel not in ordered_channels:
                default_channel = next(iter(ordered_channels), "default")
            payout_methods[method] = {
                "default_channel": default_channel,
                "channels": ordered_channels,
            }
        final[(currency, country)] = {
            "$schema": "https://json-schema.org/draft/2020-12/schema",
            "title": f"{country.title()} ({currency})",
            "currency": currency,
            "country": country,
            "payout_methods": payout_methods,
        }
    return final


def write_schemas(schemas: Dict[Tuple[str, str], Dict[str, Any]], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    for (currency, country), schema in tqdm(schemas.items(), desc="Writing schemas"):
        schema_path = out_dir / f"schema_{currency}_{country}.json"
        with schema_path.open("w", encoding="utf-8") as handle:
            json.dump(schema, handle, indent=2, ensure_ascii=True)


def validate_columns(df: pd.DataFrame) -> None:
    missing = EXPECTED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(f"Workbook is missing expected columns: {sorted(missing)}")


def ingest(workbook_path: Path) -> None:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    build_color_state_map(workbook_path)
    excel_workbook = load_workbook(workbook_path)
    color_workbook = None
    try:
        color_workbook = openpyxl_load_workbook(workbook_path, data_only=True)
    except ValueError as exc:
        message = str(exc)
        if "could not read stylesheet" not in message and "Max value is" not in message:
            raise
    except Exception:
        color_workbook = None
    combined_frames: List[pd.DataFrame] = []

    for sheet_name in excel_workbook.sheet_names:
        method = infer_method(sheet_name)
        if method is None:
            continue
        sheet_df = pd.read_excel(excel_workbook, sheet_name=sheet_name, engine=excel_workbook.engine)
        worksheet = None
        if color_workbook is not None and sheet_name in color_workbook.sheetnames:
            worksheet = color_workbook[sheet_name]
        normalized = normalize_sheet(excel_workbook, sheet_name, sheet_df, worksheet, method)
        if normalized.empty:
            continue
        validate_columns(normalized)
        normalized["Notes"] = normalized["Notes"].where(normalized["Notes"].astype(bool), sheet_name)
        combined_frames.append(normalized)

    if not combined_frames:
        if color_workbook is not None:
            color_workbook.close()
        excel_workbook.close()
        raise ValueError("Workbook did not contain any data rows.")

    merged = pd.concat(combined_frames, ignore_index=True)
    if color_workbook is not None:
        color_workbook.close()
    excel_workbook.close()
    schemas = build_schemas(merged)
    write_schemas(schemas, settings.SCHEMA_DIR)
    print(f"Generated {len(schemas)} corridor schema files in {settings.SCHEMA_DIR}.")


def main(args: Iterable[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Convert validation workbook into JSON Schemas.")
    parser.add_argument(
        "--workbook",
        type=Path,
        default=settings.BASE_DIR / "data" / "Nium_Validation_Fields.xlsx",
        help="Path to the Excel workbook.",
    )
    parser.add_argument("--debug-country", type=str, default=None, help="Debug country code (e.g. AUSTRALIA).")
    parser.add_argument("--debug-currency", type=str, default=None, help="Debug currency code (e.g. AUD).")
    parser.add_argument("--debug-method", type=str, default=None, help="Debug payout method (bank, wallet, ...).")
    parser.add_argument("--debug-channel", type=str, default=None, help="Debug channel (local, wire, ...).")
    parsed = parser.parse_args(args=args)
    DEBUG.configure(
        country=parsed.debug_country,
        currency=parsed.debug_currency,
        method=parsed.debug_method,
        channel=parsed.debug_channel,
    )
    ingest(parsed.workbook)


if __name__ == "__main__":
    main()
